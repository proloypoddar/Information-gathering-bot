import requests
import openpyxl
from datetime import datetime

access_token = "EAAWabZAhqkp8BADZCr2savLjBFjVOF4CDcE0ATxwacpWPoJFM83sUatSLZAoyQp4kUutjkg5QPcHT9GWWZAIjXaf3TikhoPW09DFB6ZALC3cS2dQozuOhxn3fACpWeyZCfHajZCi0Id8rWJlYuWzl2f2ozvsZCCFHAzWj2YNDjlDHajLn0rxqF5BVgkweFbVcxYZD"

def get_event_info(link):
    # extract the event/page ID from the link
    if "events" in link:
        id = link.split("/")[-1].split("?")[0]
    else:
        id = link.split("/")[-1]

    # make the API request to get event information
    url = f"https://graph.facebook.com/v13.0/{id}"
    params = {"fields": "name,start_time,end_time", "access_token": access_token}
    response = requests.get(url, params=params).json()

    # extract event information from the API response
    try:
        name = response["name"]
    except KeyError:
        name = "N/A"

    try:
        start_time = datetime.strptime(response["start_time"], "%Y-%m-%dT%H:%M:%S%z")
    except KeyError:
        start_time = "N/A"

    try:
        end_time = datetime.strptime(response["end_time"], "%Y-%m-%dT%H:%M:%S%z")
    except KeyError:
        end_time = "N/A"

    return name, start_time, end_time


# example usage
event_links = [
    "https://www.facebook.com/events/1230380877560897",
    "https://www.facebook.com/BRACU.Robotics.Club/"
]

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Upcoming Events"

row_num = 1
sheet.cell(row=row_num, column=1).value = "Name"
sheet.cell(row=row_num, column=2).value = "Start Time"
sheet.cell(row=row_num, column=3).value = "End Time"

for link in event_links:
    name, start_time, end_time = get_event_info(link)
    row_num += 1
    sheet.cell(row=row_num, column=1).value = name
    sheet.cell(row=row_num, column=2).value = start_time
    sheet.cell(row=row_num, column=3).value = end_time

wb.save("upcoming_events.xlsx")
